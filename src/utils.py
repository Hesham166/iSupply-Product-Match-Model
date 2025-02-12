import os
import logging
import torch
import re
import pandas as pd
from openpyxl import load_workbook


def setup_logging(log_file=None, level=logging.INFO):
    """ Configure logging to console and optionally to a file. """
    logger = logging.getLogger()
    logger.setLevel(level)
    formatter = logging.Formatter(
        '[%(asctime)s] %(levelname)s - %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )
    if logger.hasHandlers():
        logger.handlers.clear()
    ch = logging.StreamHandler()
    ch.setFormatter(formatter)
    logger.addHandler(ch)
    if log_file:
        os.makedirs(os.path.dirname(log_file), exist_ok=True)
        fh = logging.FileHandler(log_file)
        fh.setFormatter(formatter)
        logger.addHandler(fh)


def save_checkpoint(state, filename='checkpoint.pth'):
    """ Save model and optimizer state. """
    torch.save(state, filename)


def load_checkpoint(filename, model, optimizer=None):
    """ Load model (and optimizer) state from a checkpoint. """
    checkpoint = torch.load(filename)
    model.load_state_dict(checkpoint['model_state_dict'])
    if optimizer is not None and 'optimizer_state_dict' in checkpoint:
        optimizer.load_state_dict(checkpoint['optimizer_state_dict'])
    return checkpoint


def build_vocab(df, columns, special_tokens=["<pad>", "<unk>"]):
    """
    Build a character-level vocabulary from the specified DataFrame columns.
    Args:
        df (DataFrame): Input data.
        columns (list): List of column names to scan.
        special_tokens (list): List of tokens to reserve.
    Returns:
        dict: Mapping from character to index.
    """
    chars = set([ch for col in columns for text in df[col] for ch in text])
    vocab = {}
    for idx, token in enumerate(special_tokens):
        vocab[token] = idx
    for idx, ch in enumerate(sorted(chars), start=len(special_tokens)):
        vocab[ch] = idx

    return vocab


def tokenize(s, vocab, max_len):
    """
    Tokenizes string.
    Args:
        s (str): String to be tokenized.
        vocab (dict): Mapping from character to index.
        max_len (int): Maximum sequence length (for padding/truncation).
    Returns:
        Tensor: tokenized string as a tensor of numbers.
    """
    s = str(s).lower()

    tokens = [vocab.get(ch, vocab.get("<unk>")) for ch in s]
    if len(tokens) < max_len:
        tokens = tokens + [vocab.get("<pad>")] * (max_len - len(tokens))
    else:
        tokens = tokens[:max_len]

    return torch.tensor(tokens, dtype=torch.long)


def add_top_k_scores_to_df(valid_df, scores, master_file_df, k=3):
    """ 
    Adds top-k predictions to `valid_df` based on similarity scores.
    
    Args:
        valid_df (pd.DataFrame): DataFrame to which predictions will be added.
        scores (list): A list where each element is a list of tuples 
            (prediction, similarity, index).
        master_file_df (pd.DataFrame): DataFrame containing at least the 'sku' column.
        k (int): Number of top predictions to consider.

    Returns:
        pd.DataFrame: Updated valid_df with top-k SKU, prediction, and similarity scores.
    """
    # Keep only the top-k scores per sample
    topk_scores = [score[:k] for score in scores]

    # Transpose the list to process each rank separately
    for i, rank in enumerate(zip(*topk_scores)):
        indices = [idx for _, _, idx in rank]
        preds = [pred for pred, _, _ in rank]
        sims = [sim for _, sim, _ in rank]

        valid_df[f'sku{i + 1}'] = master_file_df.loc[indices, 'sku'].values
        valid_df[f'pred{i + 1}'] = preds
        valid_df[f'sim{i + 1}'] = sims

    return valid_df


def save_dataframe_to_xlsx(df, file_path, sheet_name="Sheet1"):
    """
    Saves or appends a Pandas DataFrame to an Excel (.xlsx) file.
    
    If the file already exists, it appends (or replaces the specified sheet)
    without affecting the other sheets. If the file does not exist, a new file is created.
    
    Args:
        df (pd.DataFrame): The DataFrame to save.
        file_path (str): The file path where the Excel file should be saved.
        sheet_name (str): The sheet name in the Excel file. Default is "Sheet1".
        
    Returns:
        None
    """
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    
    if os.path.exists(file_path):
        try:
            book = load_workbook(file_path)
            writer = pd.ExcelWriter(file_path, engine='openpyxl')
            writer.book = book
            
            # If the sheet already exists, remove it to replace with the new one
            if sheet_name in book.sheetnames:
                print(f"Sheet '{sheet_name}' already exists and will be replaced.")
                std = book[sheet_name]
                book.remove(std)
            
            writer.sheets = {ws.title: ws for ws in book.worksheets}
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            writer.save()
            print(f"DataFrame appended successfully to {file_path} in sheet '{sheet_name}'.")
        
        except Exception as e:
            print(f"Error while appending to the workbook: {e}")
            # Fallback: Save as a new file
            df.to_excel(file_path, sheet_name=sheet_name, index=False, engine='openpyxl')
            print(f"DataFrame saved successfully to new file {file_path} in sheet '{sheet_name}'.")
    
    else:
        # If the workbook does not exist, create it
        df.to_excel(file_path, sheet_name=sheet_name, index=False, engine='openpyxl')
        print(f"DataFrame saved successfully to new file {file_path} in sheet '{sheet_name}'.")



######################## Text Preprocessing ########################

def clean_text(text: str) -> str:
    """
    Preprocess strings by:
        - adding spaces around numbers,
        - removing all non-alphanumeric characters,
        - trimming whitespaces,
        - converting all characters to lower case,
        - and for Arabic text: normalizing different forms of alef, removing tatweel and diacritics.
    """
    # Add spaces around numbers
    text = re.sub(r"(\d+)", r" \1 ", text)
    # Remove non-alphanumerics, trim whitespaces, and convert to lower case
    text = re.sub(r"(?ui)\W", " ", text)
    text = text.strip().lower()
    # Normalize different forms of alef
    text = re.sub(r"[إأآا]", "ا", text)
    # Remove tatweel (stretching character)
    text = re.sub(r"ـ", "", text)
    # Remove diacritics (tashkeel)
    text = re.sub(r"[ًٌٍَُِّْ]", "", text)
    # Normalize whitespace: replace multiple spaces with a single space and trim the ends
    text = re.sub(r'\s+', ' ', text).strip()

    return text


def clean_df(df, text_columns, price_column, drop_duplicates=True):
    """
    Clean the DataFrame by processing text columns and fixing the price column.
    
    Parameters:
        df (pd.DataFrame): The input DataFrame.
        text_columns (list): List of column names to clean.
        price_column (str): The column name that contains price data.
        drop_duplicates (bool): Whether to drop duplicate rows.
    
    Returns:
        pd.DataFrame: The cleaned DataFrame.
    """
    # Create a copy of the DataFrame to avoid modifying the original
    df = df.copy()

    # Clean specified text columns
    for col in text_columns:
        df[col] = df[col].apply(clean_text)

    # Fix the price column if it is of type object (e.g. a string with commas)
    if df[price_column].dtype == 'O':
        df[price_column] = df[price_column].apply(lambda x: float(re.sub(',', '', str(x))))
    
    # Optionally, drop duplicate rows and reset the index
    if drop_duplicates:
        df.drop_duplicates(inplace=True)
        df.reset_index(drop=True, inplace=True)
        
    return df


def save_to_csv(df, output_name):
    """
    Save the DataFrame to a CSV file.
    
    Parameters:
        df (pd.DataFrame): The DataFrame to save.
        output_name (str): The name of the output CSV file.
    """
    # Define the directory and file path
    output_dir = './data/preprocessed'
    file_path = os.path.join(output_dir, output_name)
    
    # Create the output directory if it doesn't exist
    os.makedirs(output_dir, exist_ok=True)
    
    # Save the DataFrame to a CSV file
    df.to_csv(file_path, index=False)
    print(f"Data saved to {file_path}")


def excel_to_csv_pipeline(xlsx_path, sheet_name, text_columns, price_column, output_csv, drop_duplicates=False):
    """
    Complete processing pipeline:
      1. Load a sheet from an XLSX file.
      2. Clean text columns.
      3. Fix price column.
      4. Save the DataFrame as a CSV file.

    Parameters:
        xlsx_path (str): Path to the Excel file.
        sheet_name (str or int): The sheet name or index to load.
        text_columns (list): List of column names containing text to be cleaned.
        price_column (str): The column name with price data.
        output_csv (str): The output CSV file name.
    """
    # Step 1: Load the data from the specified Excel sheet
    df = pd.read_excel(xlsx_path, sheet_name=sheet_name)
    print("Data loaded from Excel.")

    # Step 2 & 3: Clean text columns and fix the price column
    df_clean = clean_df(df, text_columns, price_column, drop_duplicates)
    print("Data cleaned.")
    if drop_duplicates:
        print("Duplicates dropped.")

    # Step 4: Save the cleaned DataFrame to CSV
    save_to_csv(df_clean, output_csv)
    print("Pipeline finished.")